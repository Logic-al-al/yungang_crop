from __future__ import annotations

import argparse
import csv
import re
from dataclasses import dataclass, field
from pathlib import Path
from typing import Iterable

import cv2
import numpy as np
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter
from PIL import Image

import crop_data_images


OUTPUT_IMAGE_RE = re.compile(r"^(?P<book>.+)\.(?P<photo_id>\d{3}[a-z]?)$")
SOURCE_PAGE_RE = re.compile(r"_页面_(?P<page>\d+)$")
HEADERS = ["序号", "书籍名称", "照片编号", "图片注释"]
IGNORE_TOKENS = ("云冈", "雲冈", "雲网", "山西", "大同")


@dataclass(frozen=True)
class OutputPhoto:
    path: Path
    folder: str
    book: str
    photo_id: str
    page: str
    letter: str


@dataclass
class PageOCR:
    source_page: Path
    caption_crops: list[Path]
    raw_text: str
    parsed_by_letter: dict[str, str] = field(default_factory=dict)
    needs_review: bool = True


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description="Build XLSX metadata tables for cropped Yungang photos."
    )
    parser.add_argument("--data", default="data", type=Path)
    parser.add_argument("--output", default="output", type=Path)
    parser.add_argument(
        "--review-dir",
        default=Path("output") / "annotation_review",
        type=Path,
    )
    parser.add_argument("--dry-run", action="store_true")
    parser.add_argument("--no-ocr", action="store_true")
    parser.add_argument("--overwrite", action="store_true")
    parser.add_argument(
        "--pages",
        nargs="*",
        help="Optional page ids to process, e.g. --pages 033 034.",
    )
    return parser.parse_args()


def load_ocr_engine(no_ocr: bool):
    if no_ocr:
        return None
    try:
        from rapidocr_onnxruntime import RapidOCR

        return RapidOCR()
    except Exception as exc:  # pragma: no cover - depends on local install
        print(f"OCR unavailable, continuing without OCR: {exc}")
        return None


def iter_output_photos(output_root: Path) -> list[OutputPhoto]:
    photos: list[OutputPhoto] = []
    for folder_dir in sorted(p for p in output_root.iterdir() if p.is_dir()):
        if folder_dir.name == "annotation_review":
            continue
        for path in sorted(folder_dir.glob("*.jpg")):
            match = OUTPUT_IMAGE_RE.match(path.stem)
            if not match:
                continue
            photo_id = match.group("photo_id")
            page_match = re.match(r"(?P<page>\d{3})(?P<letter>[a-z]?)$", photo_id)
            if not page_match:
                continue
            photos.append(
                OutputPhoto(
                    path=path,
                    folder=folder_dir.name,
                    book=match.group("book"),
                    photo_id=photo_id,
                    page=page_match.group("page"),
                    letter=page_match.group("letter") or "",
                )
            )
    return photos


def source_pages_by_folder(data_root: Path) -> dict[tuple[str, str], Path]:
    pages: dict[tuple[str, str], Path] = {}
    for path in sorted(data_root.rglob("*.png")):
        match = SOURCE_PAGE_RE.search(path.stem)
        if match:
            pages[(path.parent.name, match.group("page"))] = path
    return pages


def crop_box_to_tuple(box: crop_data_images.CropBox) -> tuple[int, int, int, int]:
    return box.x1, box.y1, box.x2, box.y2


def detect_source_boxes(source_page: Path) -> list[crop_data_images.CropBox]:
    with Image.open(source_page) as image:
        rgb = np.array(image.convert("RGB"))
    return crop_data_images.detect_photo_boxes(rgb, max_dim=1800, safety_pixels=4)


def clamp_region(
    region: tuple[int, int, int, int], width: int, height: int
) -> tuple[int, int, int, int] | None:
    x1, y1, x2, y2 = region
    x1 = max(0, min(width, x1))
    x2 = max(0, min(width, x2))
    y1 = max(0, min(height, y1))
    y2 = max(0, min(height, y2))
    if x2 - x1 < width * 0.08 or y2 - y1 < height * 0.04:
        return None
    return x1, y1, x2, y2


def caption_regions(
    width: int, height: int, boxes: list[crop_data_images.CropBox]
) -> list[tuple[str, tuple[int, int, int, int]]]:
    if not boxes:
        return [("lower", (0, round(height * 0.62), width, height))]

    min_x = min(box.x1 for box in boxes)
    max_x = max(box.x2 for box in boxes)
    min_y = min(box.y1 for box in boxes)
    max_y = max(box.y2 for box in boxes)
    pad_y = round(height * 0.035)
    pad_x = round(width * 0.035)

    candidates = [
        ("lower", (0, max(round(height * 0.58), max_y - pad_y), width, height)),
        ("lower_wide", (0, round(height * 0.68), width, height)),
        ("below_photos", (0, max(0, max_y - round(height * 0.08)), width, height)),
        ("right_side", (max_x - pad_x, min_y, width, height)),
        ("left_side", (0, min_y, min_x + pad_x, height)),
    ]

    regions: list[tuple[str, tuple[int, int, int, int]]] = []
    seen: set[tuple[int, int, int, int]] = set()
    for label, region in candidates:
        clamped = clamp_region(region, width, height)
        if clamped and clamped not in seen:
            seen.add(clamped)
            regions.append((label, clamped))
    return regions


def save_caption_crops(
    source_page: Path,
    review_dir: Path,
    folder: str,
    regions: list[tuple[str, tuple[int, int, int, int]]],
    dry_run: bool,
) -> list[Path]:
    if dry_run:
        return [review_dir / "caption_crops" / folder / f"{source_page.stem}_{label}.jpg" for label, _ in regions]

    out_dir = review_dir / "caption_crops" / folder
    out_dir.mkdir(parents=True, exist_ok=True)
    written: list[Path] = []
    with Image.open(source_page) as image:
        image = image.convert("RGB")
        for label, region in regions:
            crop = image.crop(region)
            out_path = out_dir / f"{source_page.stem}_{label}.jpg"
            crop.save(out_path, format="JPEG", quality=95, subsampling=0)
            written.append(out_path)
    return written


def normalize_ocr_result(result) -> list[tuple[str, float]]:
    if not result:
        return []
    lines: list[tuple[str, float]] = []
    for item in result:
        if len(item) >= 3:
            text = str(item[1]).strip()
            try:
                score = float(item[2])
            except Exception:
                score = 0.0
            if text:
                lines.append((text, score))
    return lines


def prepare_for_ocr(image: Image.Image) -> list[np.ndarray]:
    rgb = np.array(image.convert("RGB"))
    resized = cv2.resize(rgb, None, fx=1.6, fy=1.6, interpolation=cv2.INTER_CUBIC)
    return [resized]


def run_ocr_on_regions(
    ocr,
    source_page: Path,
    regions: list[tuple[str, tuple[int, int, int, int]]],
) -> str:
    if ocr is None:
        return ""

    text_parts: list[str] = []
    seen: set[str] = set()
    with Image.open(source_page) as image:
        image = image.convert("RGB")
        for label, region in regions:
            crop = image.crop(region)
            for variant in prepare_for_ocr(crop):
                result, _ = ocr(variant)
                for text, score in normalize_ocr_result(result):
                    cleaned = text.strip()
                    if not cleaned or cleaned in seen:
                        continue
                    seen.add(cleaned)
                    text_parts.append(f"[{label}] {cleaned} ({score:.2f})")
    return "\n".join(text_parts)


def clean_text(text: str) -> str:
    replacements = {
        "雲": "云",
        "网": "冈",
        "岡": "冈",
        "窟 ": "窟",
        "（": "(",
        "）": ")",
        "１": "1",
        "２": "2",
        "一": "一",
    }
    for old, new in replacements.items():
        text = text.replace(old, new)
    text = re.sub(r"\s+", "", text)
    return text


def strip_ignored(text: str) -> str:
    for token in IGNORE_TOKENS:
        text = text.replace(token, "")
    text = re.sub(r"^\(?[12]\)?", "", text)
    text = re.sub(r"^[、，:：.\-]+", "", text)
    return text.strip()


def extract_ocr_text_lines(raw_text: str) -> list[str]:
    lines: list[str] = []
    for raw_line in raw_text.splitlines():
        line = re.sub(r"^\[[^\]]+\]\s*", "", raw_line)
        line = re.sub(r"\s*\([0-9.]+\)$", "", line)
        line = clean_text(line)
        if line:
            lines.append(line)
    return lines


def parse_annotations(raw_text: str) -> dict[str, str]:
    lines = extract_ocr_text_lines(raw_text)
    joined = "".join(lines)

    cave = ""
    cave_match = re.search(r"第[一二三四五六七八九十百0-9]+窟", joined)
    if cave_match:
        cave = cave_match.group(0)

    by_letter: dict[str, str] = {}
    for number, letter in (("1", "a"), ("2", "b")):
        numbered = [
            line
            for line in lines
            if re.search(rf"[\(（]?{number}[\)）]", line)
        ]
        if numbered:
            content = strip_ignored(numbered[-1])
            if cave and not content.startswith(cave):
                content = f"{cave} {content}"
            by_letter[letter] = content.strip()

    if not by_letter and cave:
        content_candidates = [
            strip_ignored(line)
            for line in lines
            if "第" not in line and not any(token in line for token in IGNORE_TOKENS)
        ]
        content_candidates = [text for text in content_candidates if text]
        if content_candidates:
            by_letter[""] = f"{cave} {content_candidates[-1]}".strip()
        else:
            by_letter[""] = cave

    return by_letter


def build_page_ocr(
    source_page: Path,
    folder: str,
    review_dir: Path,
    ocr,
    dry_run: bool,
) -> PageOCR:
    boxes = detect_source_boxes(source_page)
    with Image.open(source_page) as image:
        width, height = image.size
    regions = caption_regions(width, height, boxes)
    crop_paths = save_caption_crops(source_page, review_dir, folder, regions, dry_run)
    raw_text = run_ocr_on_regions(ocr, source_page, regions)
    parsed = parse_annotations(raw_text)
    return PageOCR(
        source_page=source_page,
        caption_crops=crop_paths,
        raw_text=raw_text,
        parsed_by_letter=parsed,
        needs_review=not bool(parsed),
    )


def ensure_workbook(path: Path) -> Workbook:
    if path.exists():
        return load_workbook(path)
    workbook = Workbook()
    workbook.active.title = "Sheet1"
    return workbook


def write_xlsx(path: Path, photos: list[OutputPhoto], page_ocr: dict[tuple[str, str], PageOCR]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    workbook = ensure_workbook(path)
    sheet = workbook.active
    sheet.delete_rows(1, sheet.max_row)
    sheet.append(HEADERS)
    for cell in sheet[1]:
        cell.font = Font(bold=True)

    for idx, photo in enumerate(photos, start=1):
        ocr_info = page_ocr.get((photo.folder, photo.page))
        annotation = ""
        if ocr_info:
            annotation = ocr_info.parsed_by_letter.get(photo.letter, "")
            if not annotation and photo.letter == "":
                annotation = ocr_info.parsed_by_letter.get("", "")
        sheet.append([idx, photo.book, photo.photo_id, annotation])

    widths = [8, 34, 14, 40]
    for idx, width in enumerate(widths, start=1):
        sheet.column_dimensions[get_column_letter(idx)].width = width
    workbook.save(path)


def write_review_csv(
    path: Path,
    photos: list[OutputPhoto],
    page_ocr: dict[tuple[str, str], PageOCR],
) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    with path.open("w", encoding="utf-8-sig", newline="") as file:
        writer = csv.writer(file)
        writer.writerow(
            [
                "来源页",
                "照片编号",
                "OCR原始文本",
                "自动解析后的图片注释",
                "是否需要人工确认",
                "题注裁片",
            ]
        )
        for photo in photos:
            ocr_info = page_ocr.get((photo.folder, photo.page))
            annotation = ""
            raw_text = ""
            crops = ""
            needs_review = "是"
            if ocr_info:
                raw_text = ocr_info.raw_text
                annotation = ocr_info.parsed_by_letter.get(photo.letter, "")
                if not annotation and photo.letter == "":
                    annotation = ocr_info.parsed_by_letter.get("", "")
                crops = ";".join(str(crop_path) for crop_path in ocr_info.caption_crops)
                needs_review = "否" if annotation else "是"
            writer.writerow(
                [
                    str(ocr_info.source_page if ocr_info else ""),
                    photo.photo_id,
                    raw_text,
                    annotation,
                    needs_review,
                    crops,
                ]
            )


def grouped_by_folder(photos: Iterable[OutputPhoto]) -> dict[str, list[OutputPhoto]]:
    grouped: dict[str, list[OutputPhoto]] = {}
    for photo in photos:
        grouped.setdefault(photo.folder, []).append(photo)
    return grouped


def main() -> None:
    args = parse_args()
    data_root = args.data.resolve()
    output_root = args.output.resolve()
    review_dir = args.review_dir.resolve()

    photos = iter_output_photos(output_root)
    if args.pages:
        selected_pages = {page.zfill(3) for page in args.pages}
        photos = [photo for photo in photos if photo.page in selected_pages]
    pages = source_pages_by_folder(data_root)
    missing_pages = [
        photo for photo in photos if (photo.folder, photo.page) not in pages
    ]
    if missing_pages:
        missing = ", ".join(photo.photo_id for photo in missing_pages[:10])
        raise SystemExit(f"Missing source pages for output photos: {missing}")

    duplicate_ids = sorted(
        {
            (photo.folder, photo.photo_id)
            for photo in photos
            if sum(
                1
                for other in photos
                if other.folder == photo.folder and other.photo_id == photo.photo_id
            )
            > 1
        }
    )
    if duplicate_ids:
        raise SystemExit(f"Duplicate photo ids: {duplicate_ids[:10]}")

    ocr = load_ocr_engine(args.no_ocr)
    needed_pages = sorted({(photo.folder, photo.page) for photo in photos})
    page_ocr: dict[tuple[str, str], PageOCR] = {}
    for key in needed_pages:
        folder, page = key
        page_ocr[key] = build_page_ocr(
            pages[key],
            folder,
            review_dir,
            ocr,
            dry_run=args.dry_run,
        )
        print(
            f"{folder}/{page}: "
            f"parsed={page_ocr[key].parsed_by_letter} "
            f"raw_lines={len(page_ocr[key].raw_text.splitlines())}"
        )

    if not args.dry_run:
        all_review_path = review_dir / "annotation_review.csv"
        write_review_csv(all_review_path, photos, page_ocr)
        for folder, folder_photos in grouped_by_folder(photos).items():
            xlsx_path = output_root / folder / f"云冈老照片信息收集表_{folder}.xlsx"
            if xlsx_path.exists() and not args.overwrite:
                raise FileExistsError(
                    f"{xlsx_path} already exists; pass --overwrite to update it"
                )
            write_xlsx(xlsx_path, folder_photos, page_ocr)

    grouped = grouped_by_folder(photos)
    total_with_annotation = 0
    for photo in photos:
        info = page_ocr.get((photo.folder, photo.page))
        if not info:
            continue
        if info.parsed_by_letter.get(photo.letter) or (
            photo.letter == "" and info.parsed_by_letter.get("")
        ):
            total_with_annotation += 1
    print(
        f"records={len(photos)} "
        f"folders={{{', '.join(f'{k}: {len(v)}' for k, v in grouped.items())}}} "
        f"annotated={total_with_annotation}"
    )


if __name__ == "__main__":
    main()
